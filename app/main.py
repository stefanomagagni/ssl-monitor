from fastapi import FastAPI
from fastapi.responses import HTMLResponse, PlainTextResponse, FileResponse
from .checker import check_domains
from .notifier import notify
import io
import csv
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
import datetime
import os

app = FastAPI()

# 🔥 ICON MAPPING (uguale per web & Excel)
PROTOCOL_ICONS = {
    "tls_modern": "🟢",
    "tls_legacy": "🟠",
    "ssl_obsolete": "🔴",
    "tcp_open_not_tls": "⚫",
    "timeout": "🕓",
    "refused": "🚫",
    "no_tls": "⚪",
    None: "⚪",
}

# 🔥 EXCEL ROW COLORS (hex)
ROW_COLORS = {
    "tls_modern": "C6EFCE",        # green
    "tls_legacy": "FFF2CC",        # yellow
    "ssl_obsolete": "F4CCCC",      # red
    "tcp_open_not_tls": "D9D9D9",  # gray
    "timeout": "FCE5CD",           # orange
    "refused": "EA9999",           # dark red
    "no_tls": "EDEDED",            # light gray
    None: "FFFFFF"
}

def sort_results(results):
    def key_fn(r):
        if "error" in r:
            return (1, 999999)
        return (0, r.get("days_left", 999999))
    return sorted(results, key=key_fn)


@app.get("/", response_class=HTMLResponse)
def dashboard():
    results = sort_results(check_domains())
    notify(results)

    html = """
    <html>
    <head>
        <title>SSL Monitor</title>
        <style>
            body {
                background-image: url('https://images.unsplash.com/photo-1518770660439-4636190af475?auto=format&fit=crop&w=1400&q=80');
                background-size: cover; background-position: center;
                background-attachment: fixed;
                color: white; font-family: Arial; text-align: center; margin:0;
            }
            header { background: rgba(0,0,0,0.65); padding:20px; }
            header img { max-height:80px; }
            h1 { font-size:2.7em; margin-top:10px; }
            .actions button {
                background:#1976d2; padding:8px 16px; margin:4px; border-radius:6px;
                border:none; color:white; cursor:pointer; font-size:0.95em;
            }
            .actions button:hover { background:#12589a; }
            .tooltip { position:relative; cursor:help; }
            .tooltip span {
                visibility:hidden; opacity:0; width:260px;
                background:black; color:white; padding:8px;
                border-radius:5px; text-align:left;
                position:absolute; left:50%; transform:translateX(-50%);
                bottom:130%; transition:0.3s;
            }
            .tooltip:hover span { visibility:visible; opacity:1; }

            table {
                width:94%; margin:10px auto 40px auto; border-collapse:collapse;
                background:rgba(0, 0, 0, 0.7); border-radius:12px; overflow:hidden;
                box-shadow:0 5px 18px rgba(0,0,0,0.55);
            }
            th { background:rgba(255,255,255,0.18); padding:14px; font-size:1.05em; }
            td { padding:12px; word-break:break-word; }
            tr:nth-child(even){ background:rgba(255,255,255,0.08); }
            tr:hover { background:rgba(255,255,255,0.18); }

            .error { color:#ff8080; font-weight:bold; }
            .issuer { font-size:0.85em; color:#e0e0e0; }
            .san { font-size:0.75em; color:#ccc; }

            .legend-container {
                margin-top:5px; margin-bottom:10px; background:rgba(0,0,0,0.55);
                display:inline-block; padding:8px 18px; border-radius:10px; font-size:1.05em;
            }
            .legend-item { margin:0 14px; display:inline-block; }
        </style>
    </head>
    <body>
        <header>
            <img src="https://raw.githubusercontent.com/stefanomagagni/ssl-monitor/main/app/logo_deda.png">
            <h1>SSL Monitor</h1>
            <div class="actions">
                <button onclick="window.location.href='/export'">Esporta CSV</button>
                <button onclick="window.location.href='/export_xlsx'">Esporta Excel</button>
            </div>
        </header>

        <div class="legend-container">
            <span class="legend-item tooltip">🟢 TLS moderno<span>Supporta TLS1.2 / TLS1.3</span></span>
            <span class="legend-item tooltip">🟠 TLS legacy<span>Protocollo TLS datato</span></span>
            <span class="legend-item tooltip">🔴 SSL obsoleto<span>SSlv2/3 non sicuro</span></span>
            <span class="legend-item tooltip">⚫ No TLS<span>Porta aperta ma nessun SSL/TLS</span></span>
            <span class="legend-item tooltip">🚫 Rifiutata<span>Connessione negata</span></span>
            <span class="legend-item tooltip">🕓 Timeout<span>Nessuna risposta</span></span>
        </div>

        <table>
            <tr>
                <th>Service</th><th>Domain/IP</th><th>Port</th><th>Protocol</th>
                <th>Expires</th><th>Days Left</th><th>Issuer</th><th>SAN</th><th>Chain</th>
            </tr>
    """

    for r in results:
        if "error" in r:
            icon = PROTOCOL_ICONS.get(r.get("protocol"))
            html += f"""
            <tr>
                <td>{r.get('service')}</td><td>{r['domain']}</td><td>{r.get('port','')}</td>
                <td style='font-size:1.4em'>{icon}</td>
                <td colspan="5" class="error">Errore: {r['error']}</td>
            </tr>"""
        else:
            icon = PROTOCOL_ICONS.get(r.get("protocol"))

            if r["days_left"] <= 0: color="#ff4d4d"
            elif r.get("alert"): color="#ffcc00"
            else: color="lightgreen"

            issuer = r["issuer"]
            san_preview = ", ".join(r["san"][:2])
            san_full = ", ".join(r["san"])
            chain_icon = "⚠" if r.get("chain_incomplete") else "✔"

            html += f"""
            <tr>
                <td>{r.get('service')}</td><td>{r['domain']}</td><td>{r['port']}</td>
                <td style='font-size:1.4em'>{icon}</td>
                <td>{r['expires']}</td>
                <td style='color:{color}; font-weight:bold;'>{r['days_left']}</td>
                <td class='issuer tooltip'>{issuer[:40]}...<span>{issuer}</span></td>
                <td class='san tooltip'>{san_preview}...<span>{san_full}</span></td>
                <td>{chain_icon}</td>
            </tr>
            """

    html += """
        </table><footer>© 2025 Deda Next – Internal SSL Monitoring Dashboard</footer></body></html>"""
    return html


# ----------------------------------------------------------------------
#       EXPORT XLSX (ICON + COLOR)
# ----------------------------------------------------------------------

@app.get("/export_xlsx")
def export_xlsx():
    results = sort_results(check_domains())

    filename = f"ssl_report_{datetime.datetime.now().strftime('%Y-%m-%d')}.xlsx"
    filepath = f"/tmp/{filename}"

    wb = Workbook()
    ws = wb.active
    ws.title = "SSL Report"

    # Freeze header
    ws.freeze_panes = "A4"

    # Legend
    ws.append(["Legenda:", "🟢 TLS moderno", "🟠 TLS legacy", "🔴 SSL obsoleto", "⚫ No TLS", "🚫 Rifiutata", "🕓 Timeout"])
    ws.append([])
    headers = ["Service", "Domain", "Port", "Protocol", "Expires", "Days Left", "Issuer", "SAN", "Chain"]
    ws.append(headers)

    # Style header
    for cell in ws[3]:
        cell.fill = PatternFill(start_color="4F81BD", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for r in results:
        proto = r.get("protocol")
        icon = PROTOCOL_ICONS.get(proto)
        san = "; ".join(r.get("san") or [])
        chain = r.get("chain", f"ERROR: {r['error']}") if "error" in r else r["chain"]

        row = [
            r.get("service",""), r.get("domain",""), r.get("port",""),
            f"{icon}", r.get("expires",""), r.get("days_left",""),
            r.get("issuer",""), san, chain
        ]
        ws.append(row)

        fill = PatternFill(start_color=ROW_COLORS.get(proto), fill_type="solid")
        for cell in ws[ws.max_row]:
            cell.fill = fill
            cell.alignment = Alignment(wrap_text=True)

    wb.save(filepath)
    return FileResponse(filepath, filename=filename,
                        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# CSV EXPORT unchanged
@app.get("/export", response_class=PlainTextResponse)
def export_csv():
    results = sort_results(check_domains())
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Service", "Domain", "Port", "Protocol", "Expires", "Days Left", "Issuer", "SAN", "Chain/Error"])

    for r in results:
        icon = PROTOCOL_ICONS.get(r.get("protocol"))
        if "error" in r:
            writer.writerow([r.get("service"), r["domain"], r.get("port"), icon, "", "", "", "", f"ERROR: {r['error']}"])
        else:
            writer.writerow([
                r.get("service"), r["domain"], r["port"], icon,
                r["expires"], r["days_left"], r["issuer"], "; ".join(r["san"]), r["chain"]
            ])

    return PlainTextResponse(output.getvalue(), media_type="text/csv",
            headers={"Content-Disposition": 'attachment; filename="ssl_report.csv"'})
