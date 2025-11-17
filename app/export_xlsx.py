import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter


def protocol_to_icon(proto):
    return {
        "tls_modern": "🟢",
        "tls_legacy": "🟠",
        "ssl_obsolete": "🔴",
        "tcp_open_not_tls": "⚫",
        "timeout": "🕓",
        "refused": "🚫",
        "no_tls": "⚪"
    }.get(proto, "⚪")


def row_color(proto):
    return {
        "tls_modern": "C6EFCE",       # green
        "tls_legacy": "FFF2CC",       # yellow
        "ssl_obsolete": "F4CCCC",     # red
        "tcp_open_not_tls": "D9D9D9", # dark gray
        "timeout": "FCE5CD",          # orange
        "refused": "EA9999",          # strong red
        "no_tls": "EDEDED"            # light gray
    }.get(proto, "FFFFFF")            # default white


def generate_xlsx(results, file_path="/tmp/ssl_report.xlsx"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SSL Report"

    headers = ["Service", "Domain/IP", "Port", "Protocol", "Expires", "Days Left", "Issuer", "SAN", "Chain"]
    ws.append(headers)

    # header formatting
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col_num, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        ws.column_dimensions[get_column_letter(col_num)].width = 25

    # rows
    for row in results:
        proto = row.get("protocol", "no_tls")
        icon = protocol_to_icon(proto)
        san_text = "; ".join(row.get("san") or []) if "san" in row else ""

        line = [
            row.get("service", ""),
            row.get("domain", ""),
            row.get("port", ""),
            f"{icon} {proto}",
            row.get("expires", "") if "error" not in row else "",
            row.get("days_left", "") if "error" not in row else "",
            row.get("issuer", "") if "error" not in row else "",
            san_text,
            row.get("chain", "") if "error" not in row else "ERROR: " + row.get("error", "")
        ]

        ws.append(line)

        # apply row background
        fill = PatternFill(start_color=row_color(proto), end_color=row_color(proto), fill_type="solid")

        for col in range(1, len(headers)+1):
            cell = ws.cell(row=ws.max_row, column=col)
            cell.fill = fill
            cell.alignment = Alignment(wrap_text=True, vertical="center")

    wb.save(file_path)
    return file_path
